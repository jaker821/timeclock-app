from tkinter import *
import tkinter as tk
import sqlite3
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import math
import dateutil.parser as dp


class ExportDataFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.bind("<Return>", self.on_enter_key)

        # Variables
        self.start_date = None
        self.end_date = None

        # Image
        try:
            self.img = tk.PhotoImage(file="resources/logo.png")
            self.img_small = self.img.subsample(2, 2)
            tk.Label(self, image=self.img_small).pack(pady=10)
        except Exception as e:
            tk.Label(self, text="Image not found").pack(pady=10)
            print(f"Error loading logo: {e}")

        # Title
        tk.Label(self, text="Export Time Logs", font=("Helvetica", 16, "bold")).pack(pady=10)

        # Start Date Label
        start_date_frm = tk.Frame(self)
        start_date_frm.pack(pady=10)
        tk.Label(start_date_frm, text="Start Date: ").grid(row=0, column=0)
        self.start_date = DateEntry(
            start_date_frm, width=12, background="darkblue", foreground="white", borderwidth=2
        )
        self.start_date.grid(row=0, column=1)

        # End Date Label
        end_date_frm = tk.Frame(self)
        end_date_frm.pack(pady=10)
        tk.Label(end_date_frm, text="End Date: ").grid(row=0, column=0)
        self.end_date = DateEntry(
            end_date_frm, width=12, background="darkblue", foreground="white", borderwidth=2
        )
        self.end_date.grid(row=0, column=1)

        # Buttons
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        self.export_btn = tk.Button(
            btn_frame,
            text="Export to Excel",
            font=("Helvetica", 12),
            command=self.handle_export
        )
        self.export_btn.grid(row=0, column=0, pady=5)
        self.back_btn = tk.Button(
            btn_frame, text="Back", font=("Helvetica", 10), command=self.back_page
        )
        self.back_btn.grid(row=1, column=0)

        self.start_date.bind("<Return>", self.on_enter_key)
        self.end_date.bind("<Return>", self.on_enter_key)

    def handle_export(self):
        start_date = self.start_date.get_date().strftime("%Y-%m-%d")
        end_date = self.end_date.get_date().strftime("%Y-%m-%d")
        audit_dict, emp_totals, date_range = self.collect_time_logs(start_date, end_date)
        self.export_to_excel(audit_dict, emp_totals, date_range)

    def collect_time_logs(self, start_date, end_date):
        import dateutil.parser as dp

        conn = sqlite3.connect("timeclock.db")
        cursor = conn.cursor()

        # Get all employees
        cursor.execute("SELECT username, id FROM users WHERE role='employee' ORDER BY id ASC")
        all_users = cursor.fetchall()
        all_usernames = [u[0] for u in all_users]
        user_ids = {u[0]: u[1] for u in all_users}

        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
        date_range = [start_dt + timedelta(days=i) for i in range((end_dt - start_dt).days + 1)]

        audit_dict = {u: {d: [] for d in date_range} for u in all_usernames}
        emp_totals = {u: {'regular': 0.0, 'overtime': 0.0} for u in all_usernames}

        # Extend to cover full weeks for OT calculation
        first_sunday = start_dt - timedelta(days=start_dt.weekday() + 1 if start_dt.weekday() < 6 else 0)
        last_saturday = end_dt + timedelta(days=(5 - end_dt.weekday()) if end_dt.weekday() < 6 else 0)

        # Fetch all time logs for extended range
        cursor.execute("""
            SELECT id, user_id, clock_in_time, clock_out_time, manual_override
            FROM time_logs
            WHERE date(clock_in_time) BETWEEN ? AND ?
            ORDER BY user_id ASC
        """, (first_sunday.strftime("%Y-%m-%d"), last_saturday.strftime("%Y-%m-%d")))
        data = cursor.fetchall()

        # Store shifts per user for weekly OT
        weekly_shifts = {u: [] for u in all_usernames}

        for time_log_id, user_id, clock_in, clock_out, mo in data:
            username = [u for u, uid in user_ids.items() if uid == user_id][0]
            clock_in_dt = dp.parse(clock_in)
            day = clock_in_dt.date()

            if clock_out is None:
                shift_str = f"{clock_in_dt.strftime('%H:%M')}-???"
                if day in audit_dict[username]:
                    audit_dict[username][day].append((shift_str, mo, False))
                continue

            clock_out_dt = dp.parse(clock_out)

            # Raw hours
            hours_worked = (clock_out_dt - clock_in_dt).total_seconds() / 3600

            # Subtract lunch breaks
            cursor.execute("SELECT SUM(duration_minutes) FROM lunch_breaks WHERE time_log_id = ?", (time_log_id,))
            result = cursor.fetchone()
            lunch_minutes = result[0] or 0
            hours_worked -= lunch_minutes / 60
            hours_worked = max(0, hours_worked)

            # Format shift string for audit (show lunch if exists)
            shift_str = f"{clock_in_dt.strftime('%H:%M')}-{clock_out_dt.strftime('%H:%M')}"
            lunch_text = f"{lunch_minutes/60:.2f} Lunch" if lunch_minutes > 0 else ""
            if day in audit_dict[username]:
                cell_display = shift_str
                if lunch_text:
                    cell_display += f"; {lunch_text}"
                audit_dict[username][day].append((cell_display, mo, False))

            # Store for weekly OT calculation
            weekly_shifts[username].append((day, hours_worked, mo, shift_str))

        # --- Weekly OT calculation ---
        for username in all_usernames:
            # Group shifts by week (Sunday start)
            weeks = {}
            for day, hours, mo, shift_str in weekly_shifts[username]:
                week_start = day - timedelta(days=day.weekday() + 1 if day.weekday() < 6 else 0)
                weeks.setdefault(week_start, []).append((day, hours, mo, shift_str))

            for week_start, week_shifts in weeks.items():
                total_week_hours = sum(h for _, h, _, _ in week_shifts)
                cumulative_hours = 0.0

                for day, hours, mo, shift_str in week_shifts:
                    reg_hours = hours
                    ot_hours = 0.0
                    if start_dt <= day <= end_dt:
                        ot_hours = min(max(0.0, cumulative_hours + hours - 40.0), hours) if total_week_hours > 40 else 0.0
                        reg_hours = hours - ot_hours

                        emp_totals[username]['regular'] += reg_hours
                        emp_totals[username]['overtime'] += ot_hours

                    cumulative_hours += hours

        conn.close()
        return audit_dict, emp_totals, date_range



    def export_to_excel(self, audit_dict, emp_totals, date_range):
        """
        Export the collected data to an Excel file.

        Features:
        - Sheet 1: Totals per employee (regular & OT hours)
        - Sheet 2: Audit log (shifts per day)
        - Highlight manual overrides (red) and OT shifts (orange)
        - Auto-size columns and apply borders, freeze headers
        """
        file_destination = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Excel File",
        )

        if not file_destination:
            return

        try:
            wb = Workbook()

            # --- Sheet 1: Totals ---
            ws_totals = wb.active
            ws_totals.title = "Totals"
            ws_totals.append(["Username", "Regular Hours", "Overtime Hours"])

            for username in sorted(emp_totals.keys()):
                totals = emp_totals[username]
                ws_totals.append([
                    username,
                    round(totals['regular'], 2),
                    round(totals['overtime'], 2)
                ])

            # --- Sheet 2: Audit Log ---
            ws_audit = wb.create_sheet("Audit Log")

            # Legend row
            ws_audit.append(["Red = Manual Override / Adjusted | Orange = OT"])
            ws_audit.row_dimensions[1].height = 20
            ws_audit.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(date_range)+1)
            legend_cell = ws_audit.cell(row=1, column=1)
            legend_cell.font = Font(bold=True, color="9C0006")
            legend_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Header row
            header = ["Username"] + [d.strftime("%Y-%m-%d") for d in date_range]
            ws_audit.append(header)

            # Add user rows
            for row_idx, username in enumerate(sorted(audit_dict.keys()), start=3):
                ws_audit.cell(row=row_idx, column=1, value=username)
                for col_idx, d in enumerate(date_range, start=2):
                    shifts = audit_dict[username].get(d, [])
                    cell_text = "; ".join([s for s, _, _ in shifts]) if shifts else ""
                    cell = ws_audit.cell(row=row_idx, column=col_idx, value=cell_text)

                    # Highlight manual override (red)
                    if any(mo and str(mo).upper() == "Y" for _, mo, _ in shifts):
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                    # Highlight OT shifts (orange)
                    if any(is_ot for _, _, is_ot in shifts):
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

            # --- Apply styling to all sheets ---
            bold_font = Font(bold=True)
            header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for ws in [ws_totals, ws_audit]:
                ws.freeze_panes = "B2"

                # Style headers
                for cell in ws[2] if ws == ws_audit else ws[1]:
                    cell.font = bold_font
                    cell.fill = header_fill
                    cell.alignment = center_align

                # Borders & column width
                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                    for cell in col:
                        cell.border = thin_border
                    ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

            wb.save(file_destination)
            tk.messagebox.showinfo("Success", f"File saved successfully to:\n{file_destination}")

        except Exception as e:
            tk.messagebox.showerror("Error", f"Error saving file: {e}")
            print(f"Error saving file: {e}")



    def on_enter_key(self, event):
        self.handle_export()


    def back_page(self):
        self.master.current_window = "admin_frame"
        self.pack_forget()
        self.master.admin_frame.pack(fill="both", expand=True)